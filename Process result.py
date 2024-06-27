import openpyxl
import pandas as pd

def process_excel(input_csv, output_excel):
    # 读取CSV文件
    df = pd.read_csv(input_csv, encoding='gbk')
    
    # 保存为Excel格式，这里直接覆盖输出文件
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
    
    # 使用openpyxl加载工作簿
    wb = openpyxl.load_workbook(output_excel)
    ws = wb.active
    
    # 遍历第一列，从第二行开始
    for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
        cell = row[0]
        value = cell.value

        # 执行公式 RIGHT("00000" & A2, 6) 的等效操作
        if value is not None:
            # 格式化为6位数字，不足的前面补0
            cell.value = f'{value:06}' if isinstance(value, int) else f'{int(value):06}'

    # 保存工作簿，覆盖原文件
    wb.save(output_excel)

# 使用函数处理Excel
input_csv_path = 'result.csv'
output_excel_path = 'result.xlsx'
process_excel(input_csv_path, output_excel_path)
