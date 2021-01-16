import random
import time
import requests    # XPath解析
from lxml import etree
import re
from openpyxl import Workbook

# 设置股票代码、时间
tcode = '377240'
sdate = '2016-03-01'
edate = '2017-03-01'

# 初始化excel函数
wb = Workbook()
st = wb.active

# 获取股票名称
renaze = re.compile(r'(?<='+tcode+'",")\w*","([\u4e00-\u9fa5]{0,})')
rename = requests.get(url=f'http://fund.eastmoney.com/js/fundcode_search.js')
name = renaze.search(rename.text)[1]

# 获取表格标题
fdurl = 'http://fund.eastmoney.com/f10/F10DataApi.aspx?type=lsjz&code=' + \
    tcode+'&page=1&per=20&sdate='+sdate+'&edate='+edate
fddata = requests.get(url=fdurl)
htmlt = etree.HTML(fddata.text)
# 通过XPath语法从页面中提取需要的数据
# 可通过Chrome浏览器中F12调试工具右键选择查看元素的XPath★★★
fdtitles = htmlt.xpath('/html/body/table/thead/tr/th')
# title = re.search(r'', fdti[0])[0]
tirow = 0
for fdti in fdtitles:
    st.cell(row=1, column=tirow + 1).value = fdti.text
    tirow += 1

# 获取基金数据
fddatas = htmlt.xpath('/html/body/text()')
pages = re.search(r'(?<=pages:)[1-9]+', fddatas[0])[0]
spantext = []
for page in range(1, int(pages)):
    spans = htmlt.xpath(
        '/html/body/table/tbody/tr/td')
    for span in spans:
        spantext.append(span.text)

# 写入数据
sx = 0
srow = 1
while sx < len(spantext):
    for scol in range(7):
        st.cell(row=srow + 1, column=scol + 1).value = spantext[sx]
        sx += 1
    srow += 1
wb.save(f'{name}.xlsx')
print('生成完毕')
